# author: Trent Florey
# 07/11/21
# interpret fantasy football data to improve
# auction draft spending this fall
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
import time
from openpyxl import Workbook

# create a webdriver and make sure it is up to date
options = Options()
options.headless = True
driver = webdriver.Chrome(ChromeDriverManager().install(), options=options,)

def main():

    # constant value for money available ($200 each for 12 teams)
    TOTAL_MONEY = 1200

    # only positions we care about are these 4 (only ones rosterd in my league)
    positions = ["QB", "RB", "WR", "TE"]

    # this information is used for web scraping the preSeason projections
    positionIterationMap = {
        "QB": 3,
        "RB": 4,
        "WR": 6,
        "TE": 3
    }

    # these are the total expected number of rostered players by position
    # group in my league
    numberRosteredMap = {
        "QB": 30,
        "RB": 52,
        "WR": 84,
        "TE": 36
    }

    # these are the years for which I want to compare predictions and results
    years = [year for year in range(2011, 2021)]

    # one list of lists called values that holds the value for each player for
    # each position ranking each year. This will first contain player names
    # with the index being their preseason projected rank by position group then
    # it will contain the points actually produced by that player then it will
    # contain the points above replacement (0 if not above replacement)
    values = [[[] for position in positions] for year in years]

    # loop through each year, each position group within each year and iteration
    # has to do with the way the website stores the rankings (only has 10 per page
    # in some cases) but we have a number we want to reach based on numberRosteredMap
    for yearIndex in range(len(years)):
        for positionIndex in range(len(positions)):
            for iteration in range(1, positionIterationMap.get(positions[positionIndex]) + 1):
                # define the url of the preseason rankings page
                url = getPreseasonUrl(years[yearIndex], positions[positionIndex], iteration)
                try:
                    driver.get(url)
                    # scrape the names from the page into the values list
                    getPlayerNames(values[yearIndex][positionIndex], numberRosteredMap.get(positions[positionIndex]))
                except:
                    continue

            # scrape the values the players now stored in the values list produced
            getPlayerPoints(values[yearIndex][positionIndex], years[yearIndex])

            # determine the n + 1 point value for each position group
            replacement = getReplacementPlayer(positions[positionIndex], years[yearIndex], numberRosteredMap.get(positions[positionIndex]) + 1)

            # use list comprehension to subtract the replacement value from each value and set to 0 if replacement > points
            values[yearIndex][positionIndex] = [round(points - replacement, 2) if (points>replacement) else 0 for points in values[yearIndex][positionIndex]] 

    # average all values into one 2d matrix of positions where index is ranking and the points are avg points above replacement 
    avgPoints = avgValues(values, positions, years)

    # sum up all points in avgPoints
    totalPoints = sumValues(avgPoints)

    # divide total points by total money to get a dollar to points ratio
    DOLLAR_TO_POINT_RATIO = round(TOTAL_MONEY / totalPoints, 2)
    
    # multiply each points value by the dollar to points ratio to get the avg value in dollars that each position ranking was worth 
    avgPoints = [[round(DOLLAR_TO_POINT_RATIO * avgPoints[positionIndex][index], 2)
        for index in range(len(avgPoints[positionIndex]))]
        for positionIndex in range(len(positions))
    ]
    
    # output the values in an easy to read excel sheet that we will reference on draft day
    exportValues(avgPoints, positions)

def getPreseasonUrl(season, position, iteration):
    # url naming convention for each position
    positionIdMap = {
        "QB": "quarterbacks",
        "RB": "runningbacks",
        "TE": "tightends",
        "WR": "widereceivers"
    }
    positionId = positionIdMap.get(position)
    if iteration > 1:
        url = "https://www.walterfootball.com/fantasy{}{}_{}.php".format(season, positionId, iteration)
    else:
        url = "https://www.walterfootball.com/fantasy{}{}.php".format(season, positionId)

    return url

def getPlayerNames(playerNames, maxLength):
    players = driver.find_elements_by_xpath("//div[@id='MainContentBlock']/ol/li/b[1]")
    length = len(players)
    index = 0
    count = len(playerNames)

    # we only want a certain number of names based on our numberRosteredMap
    # this is passed in as maxLength so we make sure playerNames does not
    # exceed that size so go until that or if there are fewer names available
    # stop when you run out of names
    while(index < length and count < maxLength):
        text = players[index].text
        # accounts for error where michael vick's name is removed
        if(text == "QB Eagles No. 7, QB, Eagles. Bye: 7."):
            text = "Michael Vick,"
        # format the text to have just the first and last name
        commaIndex = text.index(",")
        playerNames.append(text[0:commaIndex])
        index += 1
        count += 1
    
def getPlayerPoints(values, year):
    url ="https://www.pro-football-reference.com/years/{}/fantasy.htm".format(year)
    driver.get(url)
    for index in range(len(values)):
        formattedName = formatName(values[index])
        try:
            pointTotal = driver.find_element_by_xpath("//table[@id='fantasy']/tbody/tr/td[@csk='{}']/following-sibling::td[25]".format(formattedName))
            values[index] = int(pointTotal.text)
        except:
            values[index] = 0

def formatName(name):
    spaceIndex = name.index(" ")
    first = name[:spaceIndex]
    second = name[spaceIndex+1:]
    return second + "," + first

def getReplacementPlayer(position, year, n):
    positionIdMap = {
        "QB": 2,
        "RB": 3,
        "WR": 4,
        "TE": 5
    }
    url = "https://fantasydata.com/nfl/fantasy-football-leaders?position={}&season={}&seasontype=1&scope=1&subscope=1&scoringsystem=1&startweek=1&endweek=1&aggregatescope=1&range=1".format(positionIdMap.get(position), year)
    driver.get(url)
    if(n > 50):
        # button = driver.find_element_by_xpath("//div[@class='stats-grid-container'//grid-footer//div[@class='col-xs-6 col-sm-4 load-more']/a")
        try:
            button = driver.find_element_by_xpath("//div[@class='stats-grid-container']//grid-footer//div[@class='col-xs-6 col-sm-4 load-more']/a")
        except:
            driver.quit()
            return
        # button = driver.find_element_by_xpath("//a[@class='pagesize.selected'][2]")
        # button = driver.find_element_by_xpath("//div[@class='row grid-footer']/div[@class='col-xs-6 col-sm-4 load-more']/a[1]")
        driver.execute_script("arguments[0].click();", button)
        time.sleep(5)
    
    player = driver.find_element_by_xpath("//table[@id='stats_grid']/tbody/tr[{}]/td[last()]/span".format(n))
    return float(player.text)

# function that averages the points above replacement for each position group ranking over the years calculated
def avgValues(values, positions, years):
    avg = [[0 for index in range(len(values[0][positionIndex]))] for positionIndex in range(len(positions))]
    for positionIndex in range(len(positions)):
        for index in range(len(values[0][positionIndex])):
            total = 0
            for yearIndex in range(len(years)):
                total += values[yearIndex][positionIndex][index]
            avg[positionIndex][index] = total / len(years)
    return avg

def sumValues(values):
    total = 0
    for row in values:
        for val in row:
            total += val

    return total

def exportValues(avgValues, positions):
    wb = Workbook()
    fileDest = "FFAuctionBudget.xlsx"
    ws1 = wb.active
    ws1.title = 'Auction Draft Guide'

    ws1.cell(column=2,row=1,value="Auction Draft Average Value by Position Ranking")
    for i in range(1, 85):
        ws1.cell(column = 1, row = i+3, value="{}".format(i))

    for col in range(3, len(positions) + 3):
        ws1.cell(column=col, row=3, value="{}".format(positions[col-3]))
    for col in range(3, len(avgValues) + 3):
        for row in range(4, len(avgValues[col-3]) + 4):
            ws1.cell(column=col, row=row, value="${}".format(avgValues[col-3][row-4]))
    
    wb.save(fileDest)


main()
driver.quit()